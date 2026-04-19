"""Extraction de texte : PDF (PyMuPDF), DOCX (python-docx), XLSX (openpyxl), images (Gemini)."""

import io
from typing import Optional, Tuple

import fitz  # PyMuPDF
from docx import Document as DocxDocument
from openpyxl import load_workbook

from core.config import get_settings
from services import gemini_service


def extract_text_from_pdf(file_bytes: bytes) -> Tuple[str, int]:
    """Extrait le texte d'un PDF et retourne (texte, nombre de pages)."""
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    parts: list[str] = []
    for page in doc:
        parts.append(page.get_text("text"))
    text = "\n".join(parts).strip()
    return text, doc.page_count


def extract_text_from_docx(file_bytes: bytes) -> str:
    """Extrait le texte brut d'un DOCX."""
    doc = DocxDocument(io.BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs if p.text).strip()


def extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Lit les feuilles XLSX et concatène cellules en texte."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.worksheets:
        lines.append(f"## {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append("\t".join(cells))
    return "\n".join(lines).strip()


async def extract_text_from_image(
    file_bytes: bytes,
    mime_type: str,
) -> str:
    """OCR / transcription d'image via Gemini Vision."""
    prompt = (
        "Transcris tout le texte visible sur cette image. "
        "Si c'est un document, conserve la structure. Réponds uniquement avec le texte."
    )
    return await gemini_service.describe_image_bytes(file_bytes, mime_type, prompt)


async def extract_document_text(
    file_bytes: bytes,
    filename: str,
    content_type: Optional[str],
) -> Tuple[str, int]:
    """
    Route selon l'extension / type : retourne (texte, page_count estimée).
    page_count vaut 1 pour non-PDF.
    """
    name = filename.lower()
    ct = (content_type or "").lower()

    if name.endswith(".pdf") or "pdf" in ct:
        text, pages = extract_text_from_pdf(file_bytes)
        if len(text.strip()) < 50:
            # PDF scanné ou peu de texte : essai Gemini sur le fichier entier
            try:
                if get_settings().gemini_api_key:
                    text = await gemini_service.describe_pdf_bytes(
                        file_bytes,
                        "Transcris et résume le contenu textuel de ce PDF.",
                    )
            except Exception:
                pass
        return text, pages

    if name.endswith(".docx") or "wordprocessingml" in ct:
        return extract_text_from_docx(file_bytes), 1

    if name.endswith(".xlsx") or "spreadsheetml" in ct:
        return extract_text_from_xlsx(file_bytes), 1

    if name.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")) or "image" in ct:
        mime = "image/png"
        if "jpeg" in ct or name.endswith((".jpg", ".jpeg")):
            mime = "image/jpeg"
        elif "webp" in ct:
            mime = "image/webp"
        elif "gif" in ct:
            mime = "image/gif"
        text = await extract_text_from_image(file_bytes, mime)
        return text, 1

    # Par défaut : tenter PDF puis texte brut
    try:
        return extract_text_from_pdf(file_bytes)
    except Exception:
        return file_bytes.decode("utf-8", errors="ignore"), 1
